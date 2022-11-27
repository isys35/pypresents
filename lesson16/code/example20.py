# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook

# Load in the workbook
wb = load_workbook('./test.xlsx')

# Get sheet names
print(wb.get_sheet_names())